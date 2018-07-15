# -*- coding: utf-8 -*-
"""
Created on Sat Mar 31 10:47:20 2018

@author: theism
"""
import re
import os
import pandas as pd
import gen_func as gf
from datetime import datetime, date
import logging
import openpyxl
from matplotlib import pyplot as plt
import numpy as np

# ----------------  USER EDITS -------------------------------

# input directories and information
target_dir = 'C:\\Users\\theism\\Documents\\Dimagi\\Data\\new_forms'
data_regex = re.compile(r'Forms_\d\d\d.csv')

# output_dir = 'C:\\Users\\theism\\Documents\\Dimagi\\Results\\Form Submissions\\testing'
output_dir = 'C:\\Users\\theism\\Documents\\Dimagi\\Results\\New Forms'

# if want to look at one or a subset of forms
subset_only = True
subset_forms = ['[DA] Availing of ICDS Services']

# states we care about
real_state_list = ['Madhya Pradesh', 'Chhattisgarh',
                   'Andhra Pradesh', 'Bihar', 'Jharkhand',
                   'Rajasthan', 'Uttar Pradesh', 'Maharashtra'] 

## ------------- don't edit below here -----------------------------

# start logging
gf.start_logging(output_dir)
start_time = datetime.now()

# identify inputs
if subset_only:
    folder_list = subset_forms
else:
    folder_list = os.listdir(target_dir)
date_cols = ['received_on']
col_names = ['username', 'received_on', 'userID']
location_columns = ['awc_site_code', 'awc_name', 'block_name', 'district_name', 'state_name']

# because these forms have username (which is awc_site_code), but not userID, need finagling
location_file_dir = r'C:\Users\theism\Documents\Dimagi\Data\static-awc_location.csv'
location_df = pd.read_csv(location_file_dir, usecols=location_columns)
location_df['awc_site_code'] = location_df['awc_site_code'].astype(str)

# a bit hacky of an output, but going with it for now
wb = openpyxl.Workbook()
sheet = wb.active
for row in range(1, len(real_state_list)+1):
    sheet['A' + str(row+4)].value = real_state_list[row-1]
sheet.cell(row=1, column=1).value = 'Form'
sheet.cell(row=2, column=1).value = 'Form Submissions'
sheet.cell(row=3, column=1).value = 'Unique Users Submitting'
output_col = 2

os.chdir(output_dir)

for folder in folder_list:
    if os.path.isdir(os.path.join(target_dir, folder)):
        logging.info('Going through data for: %s' % folder)
        
        if folder == 'Additional Growth Monitoring':
            col_names_to_use = col_names + ['form.measure_identify | none']
        elif folder == 'Migration':
            col_names_to_use = col_names + ['form.migrate_out.confirm_migrated_out', 'form.migrate_in.confirm_migrated_in']
        else:
            col_names_to_use = col_names

        # combine all csv into one dataframe
        input_df = gf.csv_files_to_df(os.path.join(target_dir, folder), data_regex, date_cols, col_names_to_use)
        input_df['username2'] = input_df['username'].apply(lambda x: x[1:] if x[0] == '0' else x)

        # add location information for each user        
        forms_df = pd.merge(input_df, location_df, left_on='username2', right_on='awc_site_code', how='left')
        forms_df['received_on'] = pd.to_datetime(forms_df['received_on'])
        
        # filter to real states
        # logging.info('%i users unmatched to location so far' % forms_df['awc_name'].isnull().sum())
        logging.info('only getting users from real states...')
        forms_df = forms_df.loc[(forms_df['state_name'].isin(real_state_list))]
        logging.info('%i users unmatched to location still' % forms_df['awc_name'].isnull().sum())
        logging.info('------- ANALYSIS: --------')
        logging.info('%i submissions of %s form' % (forms_df.shape[0], folder))
        num_users_submitted = forms_df['awc_name'].nunique()
        logging.info('%i different users submitted this form' % num_users_submitted)
        
        # special treatment of AGMP
        if folder == 'Additional Growth Monitoring':
            logging.info('Differentiation by measure_identify == none')
            logging.info(forms_df['form.measure_identify | none'].value_counts(dropna=False))
            logging.info('Excluding all forms with measure_identify == none')
            forms_df = forms_df[forms_df['form.measure_identify | none'].isnull()]
        
        # special treatment of Migration
        if folder == 'Migration':
            logging.info('Migration info, starting with in/out value counts:')
            logging.info(forms_df['form.migrate_out.confirm_migrated_out'].value_counts(dropna=False))
            logging.info(forms_df['form.migrate_in.confirm_migrated_in'].value_counts(dropna=False))
            # combine migration in/out columns
            forms_df['migration_out'] = forms_df['form.migrate_out.confirm_migrated_out'].apply(lambda x: str('out: ' + x) if ((x == 'yes') | (x == 'no')) else x)
            forms_df['migration_in'] = forms_df['form.migrate_in.confirm_migrated_in'].apply(lambda x: str('in: ' + x) if ((x == 'yes') | (x == 'no')) else x)
            forms_df['migration_combined'] = forms_df.apply(
                    lambda row: row['migration_out']
                    if (row['migration_out'] != '---') & (
                            row['migration_out'] != np.nan)
                    else row['migration_in'], axis=1)
            logging.info('Distribution and percent of combined migration stats')
            logging.info(forms_df['migration_combined'].value_counts(dropna=False))
            logging.info(forms_df['migration_combined'].value_counts(dropna=False) * 100. / forms_df.shape[0])
            
        state_temp_df = pd.DataFrame()
        for state in real_state_list:
            temp_data = forms_df[forms_df['state_name'] == state]['awc_name'].value_counts()
            num = len(temp_data)
            mean = temp_data.mean()
            logging.info('%s: %i users submitted (%i pct); %.2f avrg forms per user submitting' % (state, num, num*100./num_users_submitted, mean))
            
            # create outputs
            sheet.cell(row=1, column=output_col).value = folder
            sheet.cell(row=2, column=output_col).value = forms_df.shape[0]
            sheet.cell(row=3, column=output_col).value = num_users_submitted
            # number users, % of users, mean, median, max, std
            sheet.cell(row=4, column=output_col).value = 'Submitting Users'
            sheet.cell(row=real_state_list.index(state)+5, column=output_col).value = num
            sheet.cell(row=4, column=output_col+1).value = '% of Submitting Users'
            sheet.cell(row=real_state_list.index(state)+5, column=output_col+1).value = num*100./num_users_submitted
            sheet.cell(row=4, column=output_col+2).value = 'Mean Forms per User'
            sheet.cell(row=real_state_list.index(state)+5, column=output_col+2).value = mean
            sheet.cell(row=4, column=output_col+3).value = 'Median Forms per User'
            sheet.cell(row=real_state_list.index(state)+5, column=output_col+3).value = temp_data.median()
            sheet.cell(row=4, column=output_col+4).value = 'Max Forms by a User'
            sheet.cell(row=real_state_list.index(state)+5, column=output_col+4).value = temp_data.max()
            sheet.cell(row=4, column=output_col+5).value = 'Std Dev Form Submissions'
            sheet.cell(row=real_state_list.index(state)+5, column=output_col+5).value = temp_data.std()
            
            if folder == 'Migration':
                logging.info('Distribution and percent of combined migration stats by STATE')
                logging.info(forms_df[forms_df['state_name'] == state]['migration_combined'].value_counts(dropna=False))
                num_forms = len(forms_df[forms_df['state_name'] == state]['migration_combined'])
                logging.info(forms_df[forms_df['state_name'] == state]['migration_combined'].value_counts(dropna=False) * 100. / num_forms)
            
            # show max users to log
        logging.info('Top submissions by user:')
        logging.info(temp_data.head(10))
        logging.info('--------------------------')
        output_col += 6
        
        # make a plot
        plot_df = forms_df.reset_index().set_index(['received_on'])
        plt.figure(figsize=(16, 6))
        for state in real_state_list:
            plt.plot(plot_df[plot_df['state_name'] == state]['awc_name'].resample('1D').count(), label=state)
        plt.legend()
        plt.title('Form Submissions by State for %s' % folder)
        plt.grid(True)
        plt.savefig(folder + '_time_plot_' + date.today().strftime('%Y-%m-%d') + '.png')
        
wb.save('output_file_' + date.today().strftime('%Y-%m-%d') + '.xlsx')
